import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

wb = openpyxl.Workbook()

# helpers
def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, size=10, color="000000", italic=False, underline=None):
    return Font(bold=bold, size=size, color=color, italic=italic, underline=underline)

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

DARK_BLUE   = "1F4E79"
MED_BLUE    = "2E75B6"
LIGHT_BLUE  = "EBF3FB"
WHITE       = "FFFFFF"
LIGHT_GREEN = "E2EFDA"
GREEN       = "C6EFCE"
GREY_TEXT   = "444444"

TITULO     = "Evaluacion de Variables y Operaciones Basicas"
ESTUDIANTE = "Katalina Dominguez"

ejercicios = [
    (
        "Variables de texto + print()",
        20, 20,
        "Perfecto. Las tres variables definidas con comillas y los tres print() funcionando correctamente. Exactamente lo que se pedia.",
    ),
    (
        "Calcular e imprimir",
        12, 20,
        "Variables bien definidas. Problema: total = 14970 declara el valor explicitamente sin usar las variables. Lo esperado era total = precio_unitario * cantidad. Ademas el print() usa suma repetida (var+var+var) en vez de multiplicacion.",
    ),
    (
        "Promedio de gastos",
        20, 20,
        "Excelente. Siguio los cuatro pasos al pie de la letra: variables de gastos, suma, division en 3 e impresion. Resultado correcto (1433.33...).",
    ),
    (
        "Renombrar variables",
        12, 20,
        "pago_por_dia es un nombre perfecto. este_mes es ambiguo (no dice cuantos dias). valores no describe nada: algo como sueldo_total o total_mes seria lo esperado. Ademas esa variable quedo sin usarse en el print().",
    ),
    (
        "Calculo completo",
        16, 20,
        "Variables bien definidas. Problema: total = 14000 declara el valor explicitamente sin utilizar las variables. Lo esperado era total = cuanto_gana * horas_trabajadas. El print() si usa las variables con multiplicacion correcta.",
    ),
]

total_obtenido = sum(e[1] for e in ejercicios)
total_maximo   = sum(e[2] for e in ejercicios)
pct = total_obtenido / total_maximo
exigencia = 0.60
if pct >= exigencia:
    nota_base = round(4.0 + 3.0 * (pct - exigencia) / (1 - exigencia), 1)
else:
    nota_base = round(1.0 + 3.0 * pct / exigencia, 1)
nota_final = nota_base  # sin decimas

mensaje = (
    "Tienes una base solida: los ejercicios donde se pedia seguir pasos estructurados los resolviste perfectamente (ej. 1 y 3). "
    "El patron a mejorar es claro: cuando necesites guardar un resultado, calcula con tus variables "
    "(total = precio_unitario * cantidad) en vez de declarar el valor directamente. "
    "Con ese ajuste tu nota sube bastante. Buen trabajo!"
)

# ─── Bienvenida ───────────────────────────────────────────────────────────────
ws_b = wb.active
ws_b.title = "Bienvenida"
ws_b.column_dimensions["A"].width = 34

ws_b["A1"] = TITULO
ws_b["A1"].font  = font(bold=True, size=13, color=WHITE)
ws_b["A1"].fill  = fill(DARK_BLUE)
ws_b["A1"].alignment = align(h="center")
ws_b.row_dimensions[1].height = 26

ws_b["A2"] = "Haz clic en tu nombre para ver tu feedback"
ws_b["A2"].font  = font(size=10, color=GREY_TEXT, italic=True)
ws_b.row_dimensions[2].height = 18

ws_b["A4"] = "Nombre del estudiante"
ws_b["A4"].font  = font(bold=True, size=10, color=WHITE)
ws_b["A4"].fill  = fill(MED_BLUE)
ws_b["A4"].alignment = align()

ws_b["A5"] = ESTUDIANTE
ws_b["A5"].font      = Font(bold=False, size=10, color="0563C1", underline="single")
ws_b["A5"].hyperlink = "#'Estudiante 1'!A1"
ws_b["A5"].alignment = align()

# ─── Estudiante 1 ─────────────────────────────────────────────────────────────
ws = wb.create_sheet("Estudiante 1")
ws.column_dimensions["A"].width = 26
ws.column_dimensions["B"].width = 10
ws.column_dimensions["C"].width = 10
ws.column_dimensions["D"].width = 52

ws.merge_cells("A1:D1")
ws["A1"] = TITULO
ws["A1"].font  = font(bold=True, size=13, color=WHITE)
ws["A1"].fill  = fill(DARK_BLUE)
ws["A1"].alignment = align(h="center")
ws.row_dimensions[1].height = 26

for col, header in zip(["A", "B", "C", "D"], ["Ejercicio", "Obtenido", "Maximo", "Observacion"]):
    c = ws[f"{col}2"]
    c.value     = header
    c.font      = font(bold=True, size=10, color=WHITE)
    c.fill      = fill(MED_BLUE)
    c.alignment = align(h="left" if col in ("A", "D") else "center")
ws.row_dimensions[2].height = 18

for i, (nombre, obtenido, maximo, obs) in enumerate(ejercicios):
    row = i + 3
    bg  = LIGHT_BLUE if i % 2 == 0 else WHITE

    ws[f"A{row}"] = nombre
    ws[f"A{row}"].font      = font(size=10)
    ws[f"A{row}"].fill      = fill(bg)
    ws[f"A{row}"].alignment = align(v="center")

    ws[f"B{row}"] = obtenido
    ws[f"B{row}"].font      = font(bold=True, size=10)
    ws[f"B{row}"].fill      = fill(bg)
    ws[f"B{row}"].alignment = align(h="center")

    ws[f"C{row}"] = maximo
    ws[f"C{row}"].font      = font(bold=True, size=10)
    ws[f"C{row}"].fill      = fill(bg)
    ws[f"C{row}"].alignment = align(h="center")

    ws[f"D{row}"] = obs
    ws[f"D{row}"].font      = font(size=9, color=GREY_TEXT)
    ws[f"D{row}"].fill      = fill(bg)
    ws[f"D{row}"].alignment = align(v="top", wrap=True)
    ws.row_dimensions[row].height = 52

spacer = len(ejercicios) + 3
ws.row_dimensions[spacer].height = 6

total_row = spacer + 1
ws.merge_cells(f"A{total_row}:C{total_row}")
ws[f"A{total_row}"] = "TOTAL"
ws[f"A{total_row}"].font      = font(bold=True, size=11, color=WHITE)
ws[f"A{total_row}"].fill      = fill(DARK_BLUE)
ws[f"A{total_row}"].alignment = align(h="center")
ws[f"D{total_row}"] = f"{total_obtenido} / {total_maximo}"
ws[f"D{total_row}"].font      = font(bold=True, size=13, color=WHITE)
ws[f"D{total_row}"].fill      = fill(DARK_BLUE)
ws[f"D{total_row}"].alignment = align(h="center")
ws.row_dimensions[total_row].height = 24

nota_row = total_row + 1
ws.merge_cells(f"A{nota_row}:C{nota_row}")
ws[f"A{nota_row}"] = "Nota"
ws[f"A{nota_row}"].font      = font(bold=True, size=11, color=DARK_BLUE)
ws[f"A{nota_row}"].fill      = fill(LIGHT_GREEN)
ws[f"A{nota_row}"].alignment = align(h="center")
ws[f"D{nota_row}"] = nota_base
ws[f"D{nota_row}"].font      = font(bold=True, size=16, color=DARK_BLUE)
ws[f"D{nota_row}"].fill      = fill(LIGHT_GREEN)
ws[f"D{nota_row}"].alignment = align(h="center")
ws.row_dimensions[nota_row].height = 28

ws.row_dimensions[nota_row + 1].height = 8

msg_row = nota_row + 2
ws.merge_cells(f"A{msg_row}:D{msg_row}")
ws[f"A{msg_row}"] = mensaje
ws[f"A{msg_row}"].font      = font(size=10, color=DARK_BLUE)
ws[f"A{msg_row}"].fill      = fill(GREEN)
ws[f"A{msg_row}"].alignment = align(v="center", wrap=True)
ws.row_dimensions[msg_row].height = 52

dec_row = msg_row + 1
ws[f"A{dec_row}"] = "Decimas"
ws[f"A{dec_row}"].font = font(bold=True)
ws[f"D{dec_row}"] = 0
ws[f"D{dec_row}"].font = font(size=11)
ws.row_dimensions[dec_row].height = 18

fin_row = dec_row + 1
ws[f"A{fin_row}"] = "Nota final"
ws[f"A{fin_row}"].font = font(bold=True, size=12)
ws[f"D{fin_row}"] = nota_final
ws[f"D{fin_row}"].font = font(bold=True, size=12)
ws.row_dimensions[fin_row].height = 20

out = "clases/reforzamiento-evaluacion/ev_kata/feedback-katalina.xlsx"
wb.save(out)
print(f"Guardado: {out}")
print(f"Total: {total_obtenido}/{total_maximo} | Nota: {nota_base} | Nota final: {nota_final}")
