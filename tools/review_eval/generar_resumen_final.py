"""
Genera los archivos de resumen final excluyendo ausentes y alonso YT.
Outputs:
  - revision/resumen_final.md
  - revision/resumen_final.csv
  - revision/resumen_final.xlsx  (para subir a Classroom)
"""

import csv
import json
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).parent.parent.parent
REVISION_DIR = BASE_DIR / "clases" / "reforzamiento-evaluacion" / "revision"

EXCLUDED = {"alonso YT"}

NAME_OVERRIDES = {
    "Estudiante Profesor Diego 1": "Alex",
}

NOTA_MIN = 2.0
NOTA_APROBACION = 4.0
NOTA_MAX = 7.0


def calcular_nota(puntaje: float, exigencia: float, puntaje_max: float = 100.0) -> float:
    corte = exigencia * puntaje_max
    if corte == 0:
        return NOTA_MIN
    if puntaje <= corte:
        nota = NOTA_MIN + (NOTA_APROBACION - NOTA_MIN) * (puntaje / corte)
    else:
        nota = NOTA_APROBACION + (NOTA_MAX - NOTA_APROBACION) * (
            (puntaje - corte) / (puntaje_max - corte)
        )
    return round(min(max(nota, NOTA_MIN), NOTA_MAX), 1)


def thin_border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)


def build_rows(data: dict) -> list[dict]:
    meta = data["metadata"]
    ej_keys = list(meta["ejercicios"].keys())
    rows = []
    for nombre_key, info in data["estudiantes"].items():
        if nombre_key in EXCLUDED:
            continue
        if not info.get("revisado"):
            continue
        total = info["total"] or 0
        if total == 0:
            continue
        display = NAME_OVERRIDES.get(nombre_key, nombre_key)
        row = {"Nombre": display}
        for ej in ej_keys:
            p = info["puntajes"].get(ej, {})
            row[ej.upper()] = p.get("obtenido") or 0
        row["Total"] = total
        row["Nota 50%"] = calcular_nota(total, 0.50)
        row["Nota 60%"] = calcular_nota(total, 0.60)
        rows.append(row)
    rows.sort(key=lambda r: r["Total"], reverse=True)
    return rows


def generar_md(rows: list[dict], meta: dict) -> None:
    ej_keys = list(meta["ejercicios"].keys())
    ej_meta = meta["ejercicios"]
    n = len(rows)
    totales = [r["Total"] for r in rows]
    notas_50 = [r["Nota 50%"] for r in rows]
    notas_60 = [r["Nota 60%"] for r in rows]
    aprobados_50 = sum(1 for x in notas_50 if x >= 4.0)
    aprobados_60 = sum(1 for x in notas_60 if x >= 4.0)

    header = (
        ["Nombre"]
        + [f"Ej{i+1} ({ej_meta[k]['maximo']}p)" for i, k in enumerate(ej_keys)]
        + ["Total", "Nota 50%", "Nota 60%"]
    )

    lines = [
        f"# Resumen Final — {meta['nombre']}",
        "",
        f"Estudiantes con entrega: {n}  |  (excluye ausentes y casos de deshonestidad académica)",
        "",
        "## Tabla de puntajes",
        "",
        "| " + " | ".join(header) + " |",
        "| " + " | ".join(["---"] * len(header)) + " |",
    ]
    for row in rows:
        vals = (
            [row["Nombre"]]
            + [str(row[k.upper()]) for k in ej_keys]
            + [str(row["Total"]), str(row["Nota 50%"]), str(row["Nota 60%"])]
        )
        lines.append("| " + " | ".join(vals) + " |")

    lines += [
        "",
        "## Estadisticas del curso",
        "",
        "| Metrica | Valor |",
        "| --- | --- |",
        f"| Estudiantes con entrega | {n} |",
        f"| Promedio puntaje | {sum(totales)/n:.1f} pts |",
        f"| Maximo | {max(totales)} pts |",
        f"| Minimo | {min(totales)} pts |",
        f"| Aprobados con exigencia 50% | {aprobados_50} / {n} ({aprobados_50/n*100:.0f}%) |",
        f"| Aprobados con exigencia 60% | {aprobados_60} / {n} ({aprobados_60/n*100:.0f}%) |",
        f"| Promedio nota (50%) | {sum(notas_50)/n:.2f} |",
        f"| Promedio nota (60%) | {sum(notas_60)/n:.2f} |",
    ]

    (REVISION_DIR / "resumen_final.md").write_text("\n".join(lines), encoding="utf-8")
    print("  OK  resumen_final.md")


def generar_csv(rows: list[dict], meta: dict) -> None:
    ej_keys = list(meta["ejercicios"].keys())
    col_names = ["Nombre"] + [k.upper() for k in ej_keys] + ["Total", "Nota 50%", "Nota 60%"]
    path = REVISION_DIR / "resumen_final.csv"
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=col_names)
        writer.writeheader()
        writer.writerows(rows)
    print("  OK  resumen_final.csv")


def generar_xlsx(rows: list[dict], meta: dict) -> None:
    ej_keys = list(meta["ejercicios"].keys())
    ej_meta = meta["ejercicios"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Notas Evaluacion 1"

    # ── Colores ───────────────────────────────────────────────────────────────
    COLOR_HEADER   = "1F4E79"   # azul oscuro
    COLOR_SUBHEAD  = "2E75B6"   # azul medio
    COLOR_APROBADO = "E2EFDA"   # verde claro
    COLOR_REPROBADO = "FFDCDC"  # rojo claro
    COLOR_ALT      = "F2F7FD"   # azul muy claro (fila alterna)
    WHITE = "FFFFFF"

    font_header = Font(name="Calibri", bold=True, color=WHITE, size=11)
    font_body   = Font(name="Calibri", size=10)
    font_bold   = Font(name="Calibri", bold=True, size=10)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left   = Alignment(horizontal="left",   vertical="center")

    # ── Título ────────────────────────────────────────────────────────────────
    total_cols = 1 + len(ej_keys) + 3  # Nombre + Ejs + Total + Nota50 + Nota60
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    title_cell = ws.cell(row=1, column=1, value=f"Evaluacion 1 — Clases 1 a 7 | {len(rows)} estudiantes con entrega")
    title_cell.font = Font(name="Calibri", bold=True, color=WHITE, size=13)
    title_cell.fill = PatternFill("solid", fgColor=COLOR_HEADER)
    title_cell.alignment = center
    ws.row_dimensions[1].height = 22

    # ── Encabezado ────────────────────────────────────────────────────────────
    headers = (
        ["Nombre"]
        + [f"Ej{i+1}\n({ej_meta[k]['maximo']}p)" for i, k in enumerate(ej_keys)]
        + ["Total", "Nota\n50%", "Nota\n60%"]
    )
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = font_header
        cell.fill = PatternFill("solid", fgColor=COLOR_SUBHEAD)
        cell.alignment = center
        cell.border = thin_border()
    ws.row_dimensions[2].height = 30

    # ── Datos ─────────────────────────────────────────────────────────────────
    for i, row in enumerate(rows):
        excel_row = i + 3
        is_even = i % 2 == 1
        nota_60 = row["Nota 60%"]
        row_fill = (
            PatternFill("solid", fgColor=COLOR_APROBADO) if nota_60 >= 4.0
            else PatternFill("solid", fgColor=COLOR_REPROBADO) if nota_60 < 3.0
            else PatternFill("solid", fgColor=COLOR_ALT if is_even else WHITE)
        )

        # Nombre
        c = ws.cell(row=excel_row, column=1, value=row["Nombre"])
        c.font = font_bold
        c.alignment = left
        c.border = thin_border()
        c.fill = row_fill

        # Ejercicios
        for j, ej in enumerate(ej_keys):
            c = ws.cell(row=excel_row, column=j + 2, value=row[ej.upper()])
            c.font = font_body
            c.alignment = center
            c.border = thin_border()
            c.fill = row_fill

        # Total
        col_total = len(ej_keys) + 2
        c = ws.cell(row=excel_row, column=col_total, value=row["Total"])
        c.font = Font(name="Calibri", bold=True, size=10)
        c.alignment = center
        c.border = thin_border()
        c.fill = row_fill

        # Nota 50%
        c = ws.cell(row=excel_row, column=col_total + 1, value=row["Nota 50%"])
        c.font = font_body
        c.alignment = center
        c.border = thin_border()
        c.fill = row_fill

        # Nota 60%
        c = ws.cell(row=excel_row, column=col_total + 2, value=row["Nota 60%"])
        c.font = Font(name="Calibri", bold=True, size=10,
                      color="375623" if nota_60 >= 4.0 else "9C0006" if nota_60 < 3.0 else "000000")
        c.alignment = center
        c.border = thin_border()
        c.fill = row_fill

        ws.row_dimensions[excel_row].height = 16

    # ── Estadísticas ──────────────────────────────────────────────────────────
    stat_row = len(rows) + 4
    ws.merge_cells(start_row=stat_row, start_column=1, end_row=stat_row, end_column=total_cols)
    ws.cell(row=stat_row, column=1, value="ESTADISTICAS DEL CURSO (estudiantes con entrega)").font = font_header
    ws.cell(row=stat_row, column=1).fill = PatternFill("solid", fgColor=COLOR_HEADER)
    ws.cell(row=stat_row, column=1).alignment = center

    n = len(rows)
    totales   = [r["Total"]    for r in rows]
    notas_50  = [r["Nota 50%"] for r in rows]
    notas_60  = [r["Nota 60%"] for r in rows]
    aprobados_50 = sum(1 for x in notas_50 if x >= 4.0)
    aprobados_60 = sum(1 for x in notas_60 if x >= 4.0)

    stats = [
        ("Estudiantes con entrega", n),
        ("Promedio puntaje", f"{sum(totales)/n:.1f} pts"),
        ("Puntaje maximo", max(totales)),
        ("Puntaje minimo", min(totales)),
        (f"Aprobados exigencia 50%", f"{aprobados_50} / {n}  ({aprobados_50/n*100:.0f}%)"),
        (f"Aprobados exigencia 60%", f"{aprobados_60} / {n}  ({aprobados_60/n*100:.0f}%)"),
        ("Promedio nota (50%)", f"{sum(notas_50)/n:.2f}"),
        ("Promedio nota (60%)", f"{sum(notas_60)/n:.2f}"),
    ]
    for k, (label, value) in enumerate(stats):
        r = stat_row + 1 + k
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=total_cols - 2)
        lc = ws.cell(row=r, column=1, value=label)
        lc.font = font_bold
        lc.alignment = left
        lc.border = thin_border()
        lc.fill = PatternFill("solid", fgColor=COLOR_ALT if k % 2 == 0 else WHITE)

        ws.merge_cells(start_row=r, start_column=total_cols - 1, end_row=r, end_column=total_cols)
        vc = ws.cell(row=r, column=total_cols - 1, value=value)
        vc.font = font_bold
        vc.alignment = center
        vc.border = thin_border()
        vc.fill = PatternFill("solid", fgColor=COLOR_ALT if k % 2 == 0 else WHITE)

    # ── Anchos de columna ─────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 32
    for col in range(2, len(ej_keys) + 2):
        ws.column_dimensions[get_column_letter(col)].width = 8
    ws.column_dimensions[get_column_letter(len(ej_keys) + 2)].width = 8   # Total
    ws.column_dimensions[get_column_letter(len(ej_keys) + 3)].width = 9   # Nota 50%
    ws.column_dimensions[get_column_letter(len(ej_keys) + 4)].width = 9   # Nota 60%

    ws.freeze_panes = "B3"

    path = REVISION_DIR / "resumen_final.xlsx"
    wb.save(path)
    print("  OK  resumen_final.xlsx")


def main() -> None:
    data = json.loads(
        (REVISION_DIR / "puntajes.json").read_text(encoding="utf-8")
    )
    rows = build_rows(data)
    print(f"\nGenerando resumen final ({len(rows)} estudiantes)...\n")
    generar_md(rows, data["metadata"])
    generar_csv(rows, data["metadata"])
    generar_xlsx(rows, data["metadata"])
    print(f"\nListo. Archivos en revision/")


if __name__ == "__main__":
    main()
