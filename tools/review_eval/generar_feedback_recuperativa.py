"""
Genera el Excel de feedback para la Evaluacion Recuperativa.
Lee: clases/reforzamiento-evaluacion/recuperativo/revision/puntajes_recuperativa.json
Output: clases/reforzamiento-evaluacion/recuperativo/revision/feedback_recuperativa.xlsx
"""

import json
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).parent.parent.parent
REVISION_DIR = BASE_DIR / "clases" / "reforzamiento-evaluacion" / "recuperativo" / "revision"

NOTA_MIN = 2.0
NOTA_APROBACION = 4.0
NOTA_MAX = 7.0

C_HEADER    = "1F4E79"
C_SUBHEAD   = "2E75B6"
C_ALT       = "EBF3FB"
C_APROBADO  = "E2EFDA"
C_MEDIO     = "FFF2CC"
C_REPROBADO = "FFDCDC"
C_MSG_AP    = "C6EFCE"
C_MSG_ME    = "FFEB9C"
C_MSG_RE    = "FFC7CE"
WHITE       = "FFFFFF"


def calcular_nota(puntaje: float, exigencia: float = 0.50, puntaje_max: float = 100.0) -> float:
    corte = exigencia * puntaje_max
    if corte == 0:
        return NOTA_MIN
    if puntaje <= corte:
        nota = NOTA_MIN + (NOTA_APROBACION - NOTA_MIN) * (puntaje / corte)
    else:
        nota = NOTA_APROBACION + (NOTA_MAX - NOTA_APROBACION) * (
            (puntaje - corte) / (puntaje_max - corte)
        )
    return round(min(max(nota, NOTA_MIN), NOTA_MAX), 1)


def nivel(nota: float) -> str:
    if nota >= 4.0:
        return "aprobado"
    if nota >= 3.0:
        return "medio"
    return "reprobado"


def mensaje_motivacional(nota: float, total: int) -> str:
    if nota >= 4.0:
        return (
            "Este resultado refleja dedicacion real y comprension solida de los conceptos. "
            "Seguir por este camino: con esta base se llega lejos. Asi se hace!"
        )
    elif total >= 30:
        return (
            "El resultado muestra que hay comprension de los conceptos fundamentales, "
            "pero falto consolidar algunos detalles clave. "
            "La distancia para aprobar es corta: con practica constante y repasando "
            "los ejercicios que costaron mas, se llega. Vale la pena el esfuerzo!"
        )
    else:
        return (
            "Los resultados de esta evaluacion no reflejan el techo de lo que se puede lograr. "
            "Programar toma tiempo y practica, es completamente normal que al inicio cueste. "
            "El profe confia en que las capacidades estan: lo que falta es seguir practicando "
            "y no rendirse. Cada error es parte del aprendizaje. Animo!"
        )


def thin():
    s = Side(style="thin", color="BBCCE4")
    return Border(left=s, right=s, top=s, bottom=s)


def fill(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=color)


def write_sheet(ws, rows: list[tuple], total: int, nota: float, titulo: str) -> None:
    nv    = nivel(nota)
    c_bg  = {"aprobado": C_APROBADO, "medio": C_MEDIO,  "reprobado": C_REPROBADO}[nv]
    c_msg = {"aprobado": C_MSG_AP,   "medio": C_MSG_ME,  "reprobado": C_MSG_RE}[nv]

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 54

    ws.merge_cells("A1:D1")
    c = ws["A1"]
    c.value     = titulo
    c.font      = Font(name="Calibri", bold=True, color=WHITE, size=13)
    c.fill      = fill(C_HEADER)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    headers = ["Ejercicio", "Obtenido", "Maximo", "Observacion"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col, value=h)
        c.font      = Font(name="Calibri", bold=True, color=WHITE, size=10)
        c.fill      = fill(C_SUBHEAD)
        c.alignment = Alignment(horizontal="center" if col > 1 else "left",
                                vertical="center", wrap_text=True)
        c.border    = thin()
    ws.row_dimensions[2].height = 18

    for i, (nombre_ej, obtenido, maximo, comentario) in enumerate(rows):
        r        = i + 3
        row_fill = fill(C_ALT if i % 2 == 0 else WHITE)

        c = ws.cell(row=r, column=1, value=nombre_ej)
        c.font      = Font(name="Calibri", size=10)
        c.fill      = row_fill
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border    = thin()

        for col, val in [(2, obtenido), (3, maximo)]:
            c = ws.cell(row=r, column=col, value=val)
            c.font      = Font(name="Calibri", bold=True, size=10)
            c.fill      = row_fill
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border    = thin()

        c = ws.cell(row=r, column=4, value=comentario)
        c.font      = Font(name="Calibri", size=9, italic=True, color="444444")
        c.fill      = row_fill
        c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        c.border    = thin()
        ws.row_dimensions[r].height = 52

    ws.row_dimensions[10].height = 6

    ws.merge_cells("A11:C11")
    c = ws["A11"]
    c.value     = "TOTAL"
    c.font      = Font(name="Calibri", bold=True, color=WHITE, size=11)
    c.fill      = fill(C_HEADER)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border    = thin()

    c = ws["D11"]
    c.value     = f"{total} / 100"
    c.font      = Font(name="Calibri", bold=True, size=13, color=WHITE)
    c.fill      = fill(C_HEADER)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border    = thin()
    ws.row_dimensions[11].height = 24

    ws.merge_cells("A12:C12")
    c = ws["A12"]
    c.value     = "Nota"
    c.font      = Font(name="Calibri", bold=True, size=11, color="1F4E79")
    c.fill      = fill(c_bg)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border    = thin()

    c = ws["D12"]
    c.value     = nota
    c.font      = Font(name="Calibri", bold=True, size=16, color="1F4E79")
    c.fill      = fill(c_bg)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border    = thin()
    ws.row_dimensions[12].height = 28

    ws.row_dimensions[13].height = 8
    ws.merge_cells("A14:D14")
    c = ws["A14"]
    c.value     = mensaje_motivacional(nota, total)
    c.font      = Font(name="Calibri", size=10, italic=True, color="1F4E79")
    c.fill      = fill(c_msg)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    c.border    = thin()
    ws.row_dimensions[14].height = 52


def write_welcome_sheet(ws, estudiantes: list[tuple], titulo: str) -> None:
    ws.column_dimensions["A"].width = 40

    c = ws["A1"]
    c.value     = titulo
    c.font      = Font(name="Calibri", bold=True, color=WHITE, size=14)
    c.fill      = fill(C_HEADER)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    c = ws["A2"]
    c.value     = "Haz clic en tu nombre para ver tu feedback"
    c.font      = Font(name="Calibri", italic=True, color="1F4E79", size=11)
    c.fill      = fill("EBF3FB")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    ws.row_dimensions[3].height = 8

    c = ws.cell(row=4, column=1, value="Nombre del estudiante")
    c.font      = Font(name="Calibri", bold=True, color=WHITE, size=10)
    c.fill      = fill(C_SUBHEAD)
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border    = thin()
    ws.row_dimensions[4].height = 18

    for i, (display_name, sheet_tab) in enumerate(sorted(estudiantes, key=lambda x: x[0].lower())):
        r        = i + 5
        row_fill = fill(C_ALT if i % 2 == 0 else WHITE)

        c = ws.cell(row=r, column=1, value=display_name)
        c.hyperlink = f"#'{sheet_tab}'!A1"
        c.font      = Font(name="Calibri", color="0563C1", underline="single",
                           bold=True, size=11)
        c.fill      = row_fill
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border    = thin()
        ws.row_dimensions[r].height = 20


def generar_feedback() -> None:
    data = json.loads((REVISION_DIR / "puntajes_recuperativa.json").read_text(encoding="utf-8"))
    meta    = data["metadata"]
    ej_keys = list(meta["ejercicios"].keys())
    titulo  = meta["nombre"]

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    estudiantes_raw = []
    for nombre_key, info in data["estudiantes"].items():
        if not info.get("revisado"):
            continue
        total      = info["total"] or 0
        nota_base  = calcular_nota(total)
        bonus      = info.get("bonus_decimas", 0)
        nota       = round(min(nota_base + bonus * 0.1, NOTA_MAX), 1)
        ej_rows = []
        for ej in ej_keys:
            ej_meta    = meta["ejercicios"][ej]
            p          = info["puntajes"].get(ej, {})
            obtenido   = p.get("obtenido") or 0
            maximo     = p.get("maximo", ej_meta["maximo"])
            comentario = p.get("comentario", "")
            ej_rows.append((ej_meta["nombre"], obtenido, maximo, comentario))
        estudiantes_raw.append((nombre_key, total, nota, ej_rows))

    estudiantes_raw.sort(key=lambda x: -x[1])
    bienvenida_data = []

    for idx, (display_name, total, nota, ej_rows) in enumerate(estudiantes_raw, start=1):
        sheet_tab = f"Estudiante {idx}"
        ws = wb.create_sheet(title=sheet_tab)
        write_sheet(ws, ej_rows, total, nota, titulo)
        bienvenida_data.append((display_name, sheet_tab))

    ws_bienvenida = wb.create_sheet(title="Bienvenida", index=0)
    write_welcome_sheet(ws_bienvenida, bienvenida_data, titulo)

    out_path = REVISION_DIR / "feedback_recuperativa.xlsx"
    wb.save(out_path)

    print(f"Excel generado ({len(estudiantes_raw)} hoja/s):\n")
    for name, total, nota, _ in sorted(estudiantes_raw, key=lambda x: -x[1]):
        print(f"  {name}: {total} pts -> nota {nota}")
    print(f"\n  -> {out_path}")


if __name__ == "__main__":
    generar_feedback()
