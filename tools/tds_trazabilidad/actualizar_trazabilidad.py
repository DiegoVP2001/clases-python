"""
Reconstruye 'clases/Ticket de Salida - Trazabilidad.xlsx' desde cero a partir de:
- respuestas_brutas.json (snapshot local de las filas del Google Sheet del Form)
- cada 'Clase NN - Tema - Ticket de Salida Respuestas.json' encontrado bajo clases/
- la nomina oficial en .claude/skills/referencia-estudiantes/lista-estudiantes.md

Fuente de verdad: este script + los JSON de respuestas correctas de cada clase.
Nunca editar el .xlsx a mano — regenerar corriendo este script.

Uso:
    python actualizar_trazabilidad.py
"""

import json
import re
import unicodedata
from collections import defaultdict
from difflib import SequenceMatcher
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).resolve().parents[2]
CLASES_DIR = BASE_DIR / "clases"
RESPUESTAS_BRUTAS = Path(__file__).resolve().parent / "respuestas_brutas.json"
NOMINA_MD = BASE_DIR / ".claude" / "skills" / "referencia-estudiantes" / "lista-estudiantes.md"
SALIDA_XLSX = CLASES_DIR / "Ticket de Salida - Trazabilidad.xlsx"

UMBRAL_MATCH_TEMA = 0.35
PREGUNTA_COMPRENSION = "Comprensión objetivo (1-5)"


def normalizar(texto: str) -> str:
    texto = texto.strip().lower()
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(c for c in texto if not unicodedata.combining(c))
    texto = re.sub(r"\s+", " ", texto)
    return texto


NO_PREGUNTADO = normalizar("No se preguntó")


def clave_orden_clase(clase) -> float:
    try:
        return float(clase)
    except (TypeError, ValueError):
        return float("inf")


def cargar_nomina() -> list[str]:
    contenido = NOMINA_MD.read_text(encoding="utf-8")
    inicio = contenido.index("## Nómina completa")
    fin = contenido.index("## Primeros nombres únicos")
    bloque = contenido[inicio:fin]
    nombres = []
    for linea in bloque.splitlines():
        m = re.match(r"\|\s*\d+\s*\|\s*(.+?)\s*\|\s*.+?\s*\|", linea)
        if m:
            nombre = m.group(1)
            nombre = re.sub(r"\s*⚠️.*$", "", nombre).strip()
            nombres.append(nombre)
    return nombres


def matchear_nombre(nombre_form: str, nomina: list[str]) -> tuple[str, bool]:
    objetivo = normalizar(nombre_form)
    for nombre_oficial in nomina:
        if normalizar(nombre_oficial) == objetivo:
            return nombre_oficial, True
    mejor = max(nomina, key=lambda n: SequenceMatcher(None, normalizar(n), objetivo).ratio())
    return mejor, False


def cargar_clases_con_ticket() -> list[dict]:
    clases = []
    for archivo in CLASES_DIR.glob("**/*Ticket de Salida Respuestas.json"):
        data = json.loads(archivo.read_text(encoding="utf-8"))
        clases.append(data)
    return clases


def matchear_tema(tema_form: str, clases: list[dict]) -> tuple[dict | None, float]:
    objetivo = normalizar(tema_form)
    mejor_clase = None
    mejor_score = 0.0
    for clase in clases:
        score = SequenceMatcher(None, normalizar(clase["tema"]), objetivo).ratio()
        if score > mejor_score:
            mejor_score = score
            mejor_clase = clase
    return mejor_clase, mejor_score


def construir_detalle(respuestas_brutas: list[dict], clases: list[dict], nomina: list[str]) -> list[dict]:
    filas = []
    for fila in respuestas_brutas:
        clase, score = matchear_tema(fila["tema"], clases)
        nombre_oficial, exacto = matchear_nombre(fila["nombre"], nomina)

        if clase is None or score < UMBRAL_MATCH_TEMA:
            filas.append({
                "fecha": fila["marca_temporal"],
                "clase": "SIN MATCH",
                "tema_clase": "—",
                "estudiante": nombre_oficial,
                "pregunta": "—",
                "respuesta": "—",
                "correcta": "—",
                "acierto": "revisar manualmente",
                "tema_escrito": fila["tema"],
            })
            continue

        for i in range(1, 5):
            respuesta = fila[f"r{i}"]
            if normalizar(respuesta) == NO_PREGUNTADO:
                continue
            correcta = clase["respuestas"].get(f"Respuestas a ticket [{i}]", "No se preguntó")
            if normalizar(correcta) == NO_PREGUNTADO:
                continue
            filas.append({
                "fecha": fila["marca_temporal"],
                "clase": clase["clase"],
                "tema_clase": clase["tema"],
                "estudiante": nombre_oficial,
                "pregunta": i,
                "respuesta": respuesta,
                "correcta": correcta,
                "acierto": "✅" if respuesta == correcta else "❌",
                "tema_escrito": fila["tema"] if not exacto else "",
            })

        comprension = fila.get("comprension", "").strip()
        if comprension and normalizar(comprension) != NO_PREGUNTADO:
            filas.append({
                "fecha": fila["marca_temporal"],
                "clase": clase["clase"],
                "tema_clase": clase["tema"],
                "estudiante": nombre_oficial,
                "pregunta": PREGUNTA_COMPRENSION,
                "respuesta": comprension,
                "correcta": "—",
                "acierto": "—",
                "tema_escrito": fila["tema"] if not exacto else "",
            })
    return filas


def hoja_detalle(wb: Workbook, filas: list[dict]):
    ws = wb.active
    ws.title = "Detalle"
    encabezados = ["Fecha", "Clase", "Tema", "Estudiante", "Pregunta", "Respuesta",
                   "Respuesta correcta", "Acierto", "Nota (match de nombre/tema)"]
    ws.append(encabezados)
    for celda in ws[1]:
        celda.font = Font(bold=True)
        celda.fill = PatternFill("solid", fgColor="D9E1F2")
    for f in filas:
        ws.append([f["fecha"], f["clase"], f["tema_clase"], f["estudiante"], f["pregunta"],
                   f["respuesta"], f["correcta"], f["acierto"], f["tema_escrito"]])
    ws.freeze_panes = "A2"
    anchos = [18, 8, 26, 32, 9, 11, 18, 9, 30]
    for i, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(i)].width = ancho


def hoja_resumen_estudiante(wb: Workbook, filas: list[dict]):
    ws = wb.create_sheet("Resumen por estudiante")
    validas = [f for f in filas if f["acierto"] in ("✅", "❌")]
    comprension_filas = [f for f in filas if f["pregunta"] == PREGUNTA_COMPRENSION]

    clases_temas = sorted(
        {(f["clase"], f["tema_clase"]) for f in validas} | {(f["clase"], f["tema_clase"]) for f in comprension_filas},
        key=lambda ct: clave_orden_clase(ct[0]),
    )
    estudiantes = sorted({f["estudiante"] for f in validas} | {f["estudiante"] for f in comprension_filas})
    n_clases = len(clases_temas)

    conteo = defaultdict(lambda: [0, 0])  # (estudiante, clase) -> [aciertos, total]
    for f in validas:
        clave = (f["estudiante"], f["clase"])
        conteo[clave][1] += 1
        if f["acierto"] == "✅":
            conteo[clave][0] += 1

    comprension_conteo = defaultdict(lambda: [0, 0])  # (estudiante, clase) -> [suma, n]
    for f in comprension_filas:
        try:
            valor = int(f["respuesta"])
        except (TypeError, ValueError):
            continue
        clave = (f["estudiante"], f["clase"])
        comprension_conteo[clave][0] += valor
        comprension_conteo[clave][1] += 1

    encabezados = (
        ["Estudiante"]
        + [f"Clase {c} - {t}" for c, t in clases_temas]
        + ["% acierto global"]
        + [f"Comprensión Clase {c} (1-5)" for c, t in clases_temas]
        + ["Comprensión promedio global"]
    )
    ws.append(encabezados)
    for celda in ws[1]:
        celda.font = Font(bold=True)
        celda.fill = PatternFill("solid", fgColor="D9E1F2")
        celda.alignment = Alignment(wrap_text=True, vertical="center")

    for estudiante in estudiantes:
        fila = [estudiante]
        aciertos_totales, total_general = 0, 0
        for clase, _ in clases_temas:
            aciertos, total = conteo.get((estudiante, clase), (0, 0))
            if total == 0:
                fila.append("—")
            else:
                fila.append(round(aciertos / total, 2))
                aciertos_totales += aciertos
                total_general += total
        fila.append(round(aciertos_totales / total_general, 2) if total_general else "—")

        suma_global, n_global = 0, 0
        for clase, _ in clases_temas:
            suma, n = comprension_conteo.get((estudiante, clase), (0, 0))
            if n == 0:
                fila.append("—")
            else:
                fila.append(round(suma / n, 2))
                suma_global += suma
                n_global += n
        fila.append(round(suma_global / n_global, 2) if n_global else "—")

        ws.append(fila)

    fila_curso = ["Promedio curso"]
    aciertos_curso, total_curso = 0, 0
    for clase, _ in clases_temas:
        aciertos_clase = sum(conteo[(e, clase)][0] for e in estudiantes if (e, clase) in conteo)
        total_clase = sum(conteo[(e, clase)][1] for e in estudiantes if (e, clase) in conteo)
        if total_clase == 0:
            fila_curso.append("—")
        else:
            fila_curso.append(round(aciertos_clase / total_clase, 2))
            aciertos_curso += aciertos_clase
            total_curso += total_clase
    fila_curso.append(round(aciertos_curso / total_curso, 2) if total_curso else "—")

    suma_curso, n_curso = 0, 0
    for clase, _ in clases_temas:
        suma_clase = sum(comprension_conteo[(e, clase)][0] for e in estudiantes if (e, clase) in comprension_conteo)
        n_clase = sum(comprension_conteo[(e, clase)][1] for e in estudiantes if (e, clase) in comprension_conteo)
        if n_clase == 0:
            fila_curso.append("—")
        else:
            fila_curso.append(round(suma_clase / n_clase, 2))
            suma_curso += suma_clase
            n_curso += n_clase
    fila_curso.append(round(suma_curso / n_curso, 2) if n_curso else "—")

    ws.append(fila_curso)
    for celda in ws[ws.max_row]:
        celda.font = Font(bold=True, italic=True)
        celda.fill = PatternFill("solid", fgColor="FCE4D6")

    col_acierto_start, col_acierto_end = 2, 1 + n_clases
    col_global_acierto = 2 + n_clases
    col_comprension_start, col_comprension_end = 3 + n_clases, 2 + 2 * n_clases
    col_comprension_global = 3 + 2 * n_clases

    for row in ws.iter_rows(min_row=2, min_col=col_acierto_start, max_col=col_global_acierto):
        for celda in row:
            if isinstance(celda.value, float):
                celda.number_format = "0%"
    for row in ws.iter_rows(min_row=2, min_col=col_comprension_start, max_col=col_comprension_global):
        for celda in row:
            if isinstance(celda.value, float):
                celda.number_format = "0.0"

    ws.freeze_panes = "B2"
    ws.column_dimensions["A"].width = 32
    for i in range(2, len(encabezados) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 16


def hoja_resumen_clase(wb: Workbook, filas: list[dict]):
    ws = wb.create_sheet("Resumen por clase")
    validas = [f for f in filas if f["acierto"] in ("✅", "❌")]
    comprension_filas = [f for f in filas if f["pregunta"] == PREGUNTA_COMPRENSION]

    claves = sorted(
        {(f["clase"], f["tema_clase"]) for f in validas} | {(f["clase"], f["tema_clase"]) for f in comprension_filas},
        key=lambda ct: clave_orden_clase(ct[0]),
    )
    preguntas = sorted({f["pregunta"] for f in validas})
    n_preguntas = len(preguntas)

    conteo = defaultdict(lambda: [0, 0])  # (clase, pregunta) -> [aciertos, total]
    for f in validas:
        clave = (f["clase"], f["pregunta"])
        conteo[clave][1] += 1
        if f["acierto"] == "✅":
            conteo[clave][0] += 1

    comprension_conteo = defaultdict(lambda: [0, 0])  # clase -> [suma, n]
    for f in comprension_filas:
        try:
            valor = int(f["respuesta"])
        except (TypeError, ValueError):
            continue
        comprension_conteo[f["clase"]][0] += valor
        comprension_conteo[f["clase"]][1] += 1

    ws.append(
        ["Clase", "Tema"] + [f"Pregunta {p}" for p in preguntas]
        + ["Comprensión promedio (1-5)", "N respuestas comprensión"]
    )
    for celda in ws[1]:
        celda.font = Font(bold=True)
        celda.fill = PatternFill("solid", fgColor="D9E1F2")

    for clase, tema in claves:
        fila = [clase, tema]
        for pregunta in preguntas:
            aciertos, total = conteo.get((clase, pregunta), (0, 0))
            fila.append(round(aciertos / total, 2) if total else "—")
        suma, n = comprension_conteo.get(clase, (0, 0))
        fila.append(round(suma / n, 2) if n else "—")
        fila.append(n)
        ws.append(fila)

    fila_general = ["—", "Promedio general"]
    for pregunta in preguntas:
        aciertos_p, total_p = 0, 0
        for clase, _ in claves:
            a, t = conteo.get((clase, pregunta), (0, 0))
            aciertos_p += a
            total_p += t
        fila_general.append(round(aciertos_p / total_p, 2) if total_p else "—")

    suma_p, n_p = 0, 0
    for clase, _ in claves:
        s, n = comprension_conteo.get(clase, (0, 0))
        suma_p += s
        n_p += n
    fila_general.append(round(suma_p / n_p, 2) if n_p else "—")
    fila_general.append(n_p)

    ws.append(fila_general)
    for celda in ws[ws.max_row]:
        celda.font = Font(bold=True, italic=True)
        celda.fill = PatternFill("solid", fgColor="FCE4D6")

    col_preguntas_start, col_preguntas_end = 3, 2 + n_preguntas
    col_comprension = 3 + n_preguntas

    for row in ws.iter_rows(min_row=2, min_col=col_preguntas_start, max_col=col_preguntas_end):
        for celda in row:
            if isinstance(celda.value, float):
                celda.number_format = "0%"
    for row in ws.iter_rows(min_row=2, min_col=col_comprension, max_col=col_comprension):
        for celda in row:
            if isinstance(celda.value, float):
                celda.number_format = "0.0"

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 26
    for i in range(3, n_preguntas + 5):
        ws.column_dimensions[get_column_letter(i)].width = 16


def main():
    respuestas_brutas = json.loads(RESPUESTAS_BRUTAS.read_text(encoding="utf-8"))
    clases = cargar_clases_con_ticket()
    nomina = cargar_nomina()

    filas = construir_detalle(respuestas_brutas, clases, nomina)

    wb = Workbook()
    hoja_detalle(wb, filas)
    hoja_resumen_estudiante(wb, filas)
    hoja_resumen_clase(wb, filas)
    wb.save(SALIDA_XLSX)

    sin_match = [f for f in filas if f["acierto"] == "revisar manualmente"]
    con_nota = [f for f in filas if f.get("tema_escrito") and f["acierto"] != "revisar manualmente"]

    print(f"Listo: {SALIDA_XLSX}")
    print(f"Filas de detalle: {len(filas)}")
    if sin_match:
        print(f"⚠️  {len(sin_match)} fila(s) sin match de clase — revisar manualmente en la hoja Detalle.")
    if con_nota:
        print(f"ℹ️  {len(con_nota)} fila(s) con tema escrito distinto al oficial (match no exacto, pero aceptado).")


if __name__ == "__main__":
    main()
