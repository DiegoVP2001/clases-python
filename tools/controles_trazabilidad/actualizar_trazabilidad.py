# -*- coding: utf-8 -*-
"""
Reconstruye 'clases/Controles - Trazabilidad.xlsx' desde cero a partir de:
- cada 'resultado_control*.json' encontrado bajo clases/ (fuente cruda escrita
  por el script de corrección de cada control, ej. generar_devolucion_control25.py)
- la nómina oficial en .claude/skills/referencia-estudiantes/lista-estudiantes.md

Misma arquitectura que tools/tds_trazabilidad/actualizar_trazabilidad.py, con
la diferencia de que la fuente cruda sale del resultado de corrección del
agente (rúbrica parcelada), no de una Google Sheet.

Fuente de verdad: este script + los resultado_control*.json de cada control.
Nunca editar el .xlsx a mano -- regenerar corriendo este script.

Uso:
    python actualizar_trazabilidad.py
"""

import json
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).resolve().parents[2]
CLASES_DIR = BASE_DIR / "clases"
NOMINA_MD = BASE_DIR / ".claude" / "skills" / "referencia-estudiantes" / "lista-estudiantes.md"
SALIDA_XLSX = CLASES_DIR / "Controles - Trazabilidad.xlsx"

FILL_HEADER = PatternFill("solid", fgColor="D9E1F2")
FONT_HEADER = Font(bold=True)
FILL_APROBADO = PatternFill("solid", fgColor="C6EFCE")
FILL_REPROBADO = PatternFill("solid", fgColor="FFC7CE")


def cargar_nomina() -> list[str]:
    contenido = NOMINA_MD.read_text(encoding="utf-8")
    inicio = contenido.index("## Nómina completa")
    fin = contenido.index("## Primeros nombres únicos")
    bloque = contenido[inicio:fin]
    nombres = []
    for linea in bloque.splitlines():
        m = re.match(r"\|\s*\d+\s*\|\s*(.+?)\s*\|\s*.+?\s*\|", linea)
        if m:
            nombre = m.group(1)
            nombre = re.sub(r"\s*⚠️.*$", "", nombre).strip()
            nombres.append(nombre)
    return nombres


def nombre_corto(nombre_completo: str, nomina: list[str]) -> str:
    """'Eduardo Pacco Ríos' calza directo contra la nómina si ya viene con
    nombre + 2 apellidos abreviados como en los resultado_control*.json (que
    usan el mismo formato corto que la revisión). Si no calza exacto, se
    devuelve tal cual -- se resuelve a mano si aparece un caso nuevo."""
    return nombre_completo


def cargar_resultados() -> list[dict]:
    """Un registro por control encontrado, con su ruta para trazabilidad."""
    controles = []
    for path in sorted(CLASES_DIR.glob("**/resultado_control*.json")):
        data = json.loads(path.read_text(encoding="utf-8"))
        data["_path"] = str(path.relative_to(BASE_DIR))
        controles.append(data)
    return controles


def construir_workbook(controles: list[dict], nomina: list[str]) -> Workbook:
    wb = Workbook()

    # --- Detalle ---
    ws = wb.active
    ws.title = "Detalle"
    headers = ["Fecha", "Clase", "Tema", "Estudiante", "Puntaje", "Máximo", "% logro", "Nota"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER

    for control in controles:
        for estudiante, r in control["estudiantes"].items():
            pct = r["puntaje"] / control["maximo"] if control["maximo"] else 0
            ws.append([
                control["fecha"], control["clase"], control["tema"], estudiante,
                r["puntaje"], control["maximo"], round(pct, 4), r["nota"],
            ])
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 28

    # --- Resumen por estudiante ---
    ws2 = wb.create_sheet("Resumen por estudiante")
    controles_ordenados = sorted(controles, key=lambda c: c["clase"])
    encabezados_control = [f"Clase {c['clase']} - {c['tema']}" for c in controles_ordenados]
    ws2.append(["Estudiante", *encabezados_control, "Promedio", "N rendidos"])
    for cell in ws2[1]:
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER

    estudiantes_vistos = set()
    for control in controles:
        estudiantes_vistos.update(control["estudiantes"].keys())
    orden = [n for n in nomina if n in estudiantes_vistos] + sorted(
        n for n in estudiantes_vistos if n not in nomina
    )

    for estudiante in orden:
        fila = [estudiante]
        notas = []
        for control in controles_ordenados:
            r = control["estudiantes"].get(estudiante)
            if r is None:
                fila.append("—")
            else:
                fila.append(r["nota"])
                notas.append(r["nota"])
        promedio = round(sum(notas) / len(notas), 2) if notas else "—"
        fila.append(promedio)
        fila.append(len(notas))
        ws2.append(fila)
        if isinstance(promedio, (int, float)):
            fill = FILL_APROBADO if promedio >= 4.0 else FILL_REPROBADO
            ws2.cell(row=ws2.max_row, column=len(fila) - 1).fill = fill

    ws2.column_dimensions["A"].width = 30
    for col in range(2, len(encabezados_control) + 3):
        ws2.column_dimensions[get_column_letter(col)].width = 22

    # --- Resumen por clase ---
    ws3 = wb.create_sheet("Resumen por clase")
    ws3.append(["Clase", "Tema", "Fecha", "N estudiantes", "Puntaje promedio", "Nota promedio",
                "% aprobación (nota ≥ 4.0)"])
    for cell in ws3[1]:
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
    for control in controles_ordenados:
        regs = list(control["estudiantes"].values())
        n = len(regs)
        puntaje_prom = round(sum(r["puntaje"] for r in regs) / n, 1) if n else "—"
        nota_prom = round(sum(r["nota"] for r in regs) / n, 2) if n else "—"
        aprob = round(100 * sum(1 for r in regs if r["nota"] >= 4.0) / n, 1) if n else "—"
        ws3.append([control["clase"], control["tema"], control["fecha"], n, puntaje_prom, nota_prom, aprob])
    for col in range(1, 8):
        ws3.column_dimensions[get_column_letter(col)].width = 20

    return wb


def main():
    nomina = cargar_nomina()
    controles = cargar_resultados()
    if not controles:
        print("No se encontró ningún resultado_control*.json bajo clases/.")
        return
    print(f"Controles encontrados: {len(controles)}")
    for c in controles:
        print(f"  - Clase {c['clase']} ({c['tema']}): {len(c['estudiantes'])} estudiantes — {c['_path']}")
    wb = construir_workbook(controles, nomina)
    wb.save(SALIDA_XLSX)
    print(f"Escrito: {SALIDA_XLSX}")


if __name__ == "__main__":
    main()
