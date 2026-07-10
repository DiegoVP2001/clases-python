import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import sys

sys.stdout = open(sys.stdout.fileno(), mode='w', encoding='utf-8', buffering=1)

OUT = r'C:/Users/diego/OneDrive/Escritorio/claude_codex/clases/claude_python/clases-python-4tomedio/clases/reforzamiento-evaluacion/revision/tabla_notas_finales.xlsx'

students = [
    (1,  'Eduardo Eleazarth Pacco Ríos',        6.5, 3, 6.8,  ''),
    (2,  'Diego Pablo Vargas Jaqui',             2.2, 1, 2.3,  ''),
    (3,  'Alex Saravia Lara',                    3.7, 1, 3.8,  '⚠️ Puntaje puede ser cuenta de prueba'),
    (4,  'Francisco Ignacio Vega Sanhueza',      6.1, 3, 6.4,  ''),
    (5,  'Felipe Alejandro Román Brito',         5.8, 3, 6.1,  ''),
    (6,  'Julián Alonso Sandoval Arriagada',     4.2, 3, 4.5,  ''),
    (7,  'Héctor Mauricio Vergara Alarcón',      6.7, 3, 7.0,  ''),
    (8,  'Benjamín Esteban Mejías González',     6.8, 3, 7.0,  'Recuperativa'),
    (9,  'Lucas Gaspar Valenzuela Donoso',       2.1, 1, 2.2,  ''),
    (10, 'Martín Natahel Sánchez Orellana',      2.7, 3, 3.0,  'Recuperativa'),
    (11, 'Benjamín Ignacio Díaz Silva',          2.1, 1, 2.2,  ''),
    (12, 'Vicente Ignacio Benítez Muñoz',        2.9, 1, 3.0,  ''),
    (13, 'Katalina Domínguez Santibáñez',        None, 0, None, 'Ausente'),
    (14, 'Maura Isabel Muñoz Gutiérrez',         3.6, 1, 3.7,  ''),
    (15, 'Vicente Narváez Fernandez',            3.3, 1, 3.4,  ''),
    (16, 'Sebastián Andrés Ulloa Cuevas',        2.9, 1, 3.0,  ''),
    (17, 'Diego Andrés Donoso Figueroa',         2.5, 0, 2.5,  ''),
    (18, 'Felipe Aravena Cárdenas',              6.6, 3, 6.9,  ''),
    (19, 'Cristóbal Alonso Muñoz Cubillos',      5.9, 3, 6.2,  ''),
    (20, 'Francisca Belén Parra Marínquez',      2.6, 0, 2.6,  ''),
    (21, 'Diego Antonio Peña Bustamante',        6.5, 3, 6.8,  ''),
    (22, 'Tomas Ignacio Diaz Calfunao',          5.4, 1, 5.5,  'Recuperativa'),
    (23, 'Santino Garcia Colombati',             5.5, 1, 5.6,  'Recuperativa'),
    (24, 'Luckas Martín Letelier Lavín',         5.8, 3, 6.1,  ''),
    (25, 'Simón Abrahams Delgado',               None, 0, None, 'Ausente'),
    (26, 'Julián Rafael Aravena Sagal',          4.0, 3, 4.3,  ''),
    (27, 'Diego Alonso Cifuentes Tessada',       2.8, 0, 2.8,  ''),
    (28, 'Paula Trinidad Inzunza Gálvez',        None, 0, None, 'Ausente'),
    (29, 'Damián Cristóbal Flores Silva',        2.6, 3, 2.9,  ''),
    (30, 'Alain Israel Moyano Molina',           3.9, 0, 3.9,  'Recuperativa'),
    (31, 'Víctor Alonso Olguín',                 6.6, 0, 6.6,  '⚠️ Sospecha de copia'),
]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Notas Finales'

# ── Estilos ──────────────────────────────────────────────────
thin = Side(style='thin', color='BBBBBB')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

header_fill = PatternFill(start_color='1F3864', end_color='1F3864', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF', size=11)

aprobado_fill  = PatternFill(start_color='D6E4BC', end_color='D6E4BC', fill_type='solid')
reprobado_fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
ausente_fill   = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
warning_fill   = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')

bold = Font(bold=True)
bold_red = Font(bold=True, color='C00000')

center = Alignment(horizontal='center', vertical='center')
left   = Alignment(horizontal='left',   vertical='center')

# ── Título ───────────────────────────────────────────────────
ws.merge_cells('A1:F1')
title_cell = ws['A1']
title_cell.value = 'Evaluación 1 — Python (Clases 1 a 7)  |  Notas finales'
title_cell.font = Font(bold=True, size=13, color='1F3864')
title_cell.alignment = center
ws.row_dimensions[1].height = 24

ws.row_dimensions[2].height = 6  # spacer

# ── Encabezados ──────────────────────────────────────────────
headers = ['N°', 'Nombre', 'Nota base', 'Décimas', 'Nota final', 'Observación']
col_widths = [5, 36, 12, 10, 12, 36]
for c, (h, w) in enumerate(zip(headers, col_widths), start=1):
    cell = ws.cell(row=3, column=c, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center
    cell.border = border
    ws.column_dimensions[cell.column_letter].width = w
ws.row_dimensions[3].height = 18

# ── Filas de datos ────────────────────────────────────────────
for row_offset, (n, name, nota_base, decimas, nota_final, obs) in enumerate(students):
    r = row_offset + 4

    is_ausente = nota_final is None
    is_warning = obs.startswith('⚠️')
    aprobado   = nota_final is not None and nota_final >= 4.0

    if is_warning:
        row_fill = warning_fill
    elif is_ausente:
        row_fill = ausente_fill
    elif aprobado:
        row_fill = aprobado_fill
    else:
        row_fill = reprobado_fill

    values = [n, name, nota_base, decimas, nota_final if nota_final else '—', obs]
    aligns = [center, left, center, center, center, left]

    for c, (val, aln) in enumerate(zip(values, aligns), start=1):
        cell = ws.cell(row=r, column=c, value=val)
        cell.fill = row_fill
        cell.alignment = aln
        cell.border = border

    # Nota final en negrita
    nota_cell = ws.cell(row=r, column=5)
    if nota_final is not None:
        nota_cell.font = Font(bold=True, color='CC0000' if not aprobado else '375623')
    else:
        nota_cell.font = Font(italic=True, color='888888')

    # Observación en rojo si es advertencia
    if is_warning:
        ws.cell(row=r, column=6).font = Font(color='C00000', italic=True)

    ws.row_dimensions[r].height = 16

# ── Leyenda ───────────────────────────────────────────────────
legend_row = len(students) + 5
ws.cell(row=legend_row, column=1).value = 'Leyenda:'
ws.cell(row=legend_row, column=1).font = Font(bold=True, color='555555')

legends = [
    ('D6E4BC', 'Aprobado (≥ 4.0)'),
    ('FCE4D6', 'Reprobado (< 4.0)'),
    ('F2F2F2', 'Ausente / Sin entrega'),
    ('FFF2CC', 'Requiere verificación'),
]
for i, (color, label) in enumerate(legends):
    ws.cell(row=legend_row + 1 + i, column=1).value = ''
    fill_cell = ws.cell(row=legend_row + 1 + i, column=2)
    fill_cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
    fill_cell.border = border
    ws.cell(row=legend_row + 1 + i, column=3).value = label
    ws.cell(row=legend_row + 1 + i, column=3).font = Font(color='555555')

wb.save(OUT)
print(f'Guardado en: {OUT}')
