"""
Actualiza puntajes de polar tv (corregidos) y agrega Julian Sandoval.
Luego actualiza feedback_estudiantes.xlsx: Estudiante 11 + nuevo Estudiante 23.
"""
import json
import openpyxl
from openpyxl.styles import Font, PatternFill
import sys, io

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

XLSX = r'C:/Users/diego/OneDrive/Escritorio/claude_codex/clases/claude_python/clases-python-4tomedio/clases/reforzamiento-evaluacion/revision/feedback_estudiantes.xlsx'
PUNTAJES = r'C:/Users/diego/OneDrive/Escritorio/claude_codex/clases/claude_python/clases-python-4tomedio/clases/reforzamiento-evaluacion/revision/puntajes.json'


def calc_nota(score):
    if score < 50:
        return round(2 + 2 * (score / 50), 1)
    else:
        return round(4 + 3 * ((score - 50) / 50), 1)


# ── 1. Actualizar puntajes.json ───────────────────────────────
with open(PUNTAJES, encoding='utf-8') as f:
    puntajes = json.load(f)

# Corregir polar tv
ptv = puntajes['estudiantes']['polar tv']['puntajes']
ptv['ej2']['obtenido'] = 10
ptv['ej2']['comentario'] = 'Título presente, sep usado, total con variable. Rúbrica: no descontar si título presente.'
ptv['ej4']['obtenido'] = 5
ptv['ej4']['comentario'] = '2 inputs float ✓. Fórmula usa precio_original (undefined, bug de tipeo). Descuento directo sin porcentaje. 1 print.'
ptv['ej5']['obtenido'] = 5
ptv['ej5']['comentario'] = 'Título ✓, 2 inputs reales float ✓. ganancia_neta pedida por input. integrantes hardcodeado. 2 prints.'
ptv['ej7']['obtenido'] = 6
ptv['ej7']['comentario'] = 'Header ✓, 3 inputs float ✓, saldo calculado ✓. total y promedio pedidos por input. 2 prints.'
puntajes['estudiantes']['polar tv']['total'] = 50

# Agregar Julian Sandoval
puntajes['estudiantes']['Julian Sandoval'] = {
    'archivo': 'clases\\reforzamiento-evaluacion\\revision\\notebooks\\Julian Sandoval rev.ipynb',
    'revisado': True,
    'puntajes': {
        'ej1': {'obtenido': 10, 'maximo': 10, 'comentario': 'nota1, nota2, nota3 en snake_case, promedio con /3, 4 prints con etiqueta. Perfecto.'},
        'ej2': {'obtenido': 10, 'maximo': 10, 'comentario': 'Título con emoji, sep usado, header y total calculado. Rúbrica: no descontar si título presente.'},
        'ej3': {'obtenido': 14, 'maximo': 15, 'comentario': 'Variable única saldo_inicial con +=/-= correcta. Variables de tracking extra no usadas en cálculos. Un print usa "deposito" en vez de "saldo_inicial" (-1). Output exacto.'},
        'ej4': {'obtenido': 0,  'maximo': 15, 'comentario': 'Sin código funcional. Celda vacía.'},
        'ej5': {'obtenido': 4,  'maximo': 15, 'comentario': 'Título ✓, 3 input() intentados (sintaxis incorrecta con 2 args). NameError por typo "total_recudado". Prints etiquetados presentes pero sin ejecución.'},
        'ej6': {'obtenido': 15, 'maximo': 15, 'comentario': 'Versión jueves (4 errores). Los 4 errores corregidos: nombre, indentación, underscore, operador. nota_final = 6.175 ✓'},
        'ej7': {'obtenido': 0,  'maximo': 20, 'comentario': 'Sin código funcional. Celda vacía.'},
    },
    'total': 53,
}

with open(PUNTAJES, 'w', encoding='utf-8') as f:
    json.dump(puntajes, f, ensure_ascii=False, indent=2)
print('✅ puntajes.json actualizado (polar tv corregido + Julian Sandoval agregado)')


# ── 2. Actualizar feedback_estudiantes.xlsx ───────────────────
wb = openpyxl.load_workbook(XLSX)

# ── 2a. Corregir Estudiante 11 (polar tv) ────────────────────
ws11 = wb['Estudiante 11']

updates_11 = {
    4: (10, 'Título presente, sep usado, total con variable. Rúbrica: no descontar si título presente.'),
    6: (5,  '2 inputs float ✓. Fórmula usa precio_original (undefined, bug de tipeo). Descuento directo sin porcentaje. 1 print.'),
    7: (5,  'Título ✓, 2 inputs reales float ✓. ganancia_neta pedida por input. integrantes hardcodeado. 2 prints.'),
    9: (6,  'Header ✓, 3 inputs float ✓, saldo calculado ✓. total y promedio pedidos por input. 2 prints.'),
}
for row, (pts, obs) in updates_11.items():
    ws11.cell(row=row, column=2).value = pts
    ws11.cell(row=row, column=4).value = obs

ws11.cell(row=11, column=4).value = '50 / 100'
ws11.cell(row=12, column=4).value = 4.0
ws11.cell(row=12, column=4).font = Font(bold=True)
ws11.cell(row=14, column=1).value = (
    'Buen trabajo en los ejercicios clave: ej6 impecable y la estructura de inputs '
    'estuvo presente. Para subir más la nota, lo que falta es consolidar la fórmula '
    'del porcentaje (ej4) y el cálculo acumulado paso a paso (ej3 y ej7). Con esta '
    'base se llega lejos. ¡Bien!'
)
ws11.cell(row=16, column=4).value = 4.3
ws11.cell(row=16, column=4).font = Font(bold=True, size=12)
print('  Estudiante 11 (polar tv) actualizado: 40 → 50 pts, nota 3.9 → 4.3')

# ── 2b. Crear Estudiante 23 (Julian Sandoval) ────────────────
julian_scores = [
    ('Promedio de notas',         10,  10, 'nota1, nota2, nota3 en snake_case, promedio con /3, 4 prints con etiqueta. Perfecto.'),
    ('Reporte feria escolar',     10,  10, 'Título con emoji, sep usado, header y total calculado. Rúbrica: no descontar si título presente.'),
    ('Simulador ahorro mensual',  14,  15, 'Variable única saldo_inicial con +=/-= correcta. Variables de tracking extra no usadas en cálculos. Un print usa "deposito" en vez de "saldo_inicial" (-1). Output exacto.'),
    ('Calculadora de descuento',   0,  15, 'Sin código funcional. Celda vacía.'),
    ('Reparto ganancias dulces',   4,  15, 'Título ✓, 3 input() intentados (sintaxis incorrecta con 2 args). NameError por typo "total_recudado". Prints etiquetados presentes pero sin ejecución.'),
    ('Depuracion 4 errores',      15,  15, 'Versión jueves (4 errores). Los 4 errores corregidos: nombre, indentación, underscore, operador. nota_final = 6.175 ✓'),
    ('Registro semanal de gastos', 0,  20, 'Sin código funcional. Celda vacía.'),
]
j_total = sum(s[1] for s in julian_scores)
j_nota_base = calc_nota(j_total)
j_decimas = 3
j_nota_final = round(min(7.0, j_nota_base + j_decimas * 0.1), 1)

ws23 = wb.copy_worksheet(wb['Estudiante 1'])
ws23.title = 'Estudiante 23'

ws23.cell(row=1, column=1).value = 'Evaluacion 1  |  Python - Clases 1 a 7'
ws23.cell(row=2, column=1).value = 'Ejercicio'
ws23.cell(row=2, column=2).value = 'Obtenido'
ws23.cell(row=2, column=3).value = 'Maximo'
ws23.cell(row=2, column=4).value = 'Observacion'

for i, (ej, obt, mx, obs) in enumerate(julian_scores, start=3):
    ws23.cell(row=i, column=1).value = ej
    ws23.cell(row=i, column=2).value = obt
    ws23.cell(row=i, column=3).value = mx
    ws23.cell(row=i, column=4).value = obs

ws23.cell(row=10, column=1).value = None
ws23.cell(row=11, column=1).value = 'TOTAL'
ws23.cell(row=11, column=4).value = f'{j_total} / 100'
ws23.cell(row=12, column=1).value = 'Nota'
ws23.cell(row=12, column=4).value = j_nota_base
ws23.cell(row=13, column=1).value = None
ws23.cell(row=14, column=1).value = (
    'Muy buen trabajo en lo que completaste: ej1, ej2 y ej6 perfectos, y el ej3 '
    'casi ideal. Falto completar ej4 y ej7, que eran los más complejos. Con la base '
    'que demostraste, esos ejercicios son el próximo paso. ¡Buen comienzo!'
)
ws23.cell(row=15, column=1).value = 'Décimas'
ws23.cell(row=15, column=4).value = j_decimas
ws23.cell(row=16, column=1).value = 'Nota final'
ws23.cell(row=16, column=4).value = j_nota_final

for r in [11, 12, 15]:
    ws23.cell(row=r, column=1).font = Font(bold=True)
ws23.cell(row=16, column=1).font = Font(bold=True, size=12)
ws23.cell(row=16, column=4).font = Font(bold=True, size=12)

print(f'  Estudiante 23 (Julian Sandoval) creado: {j_total}pts → nota_base={j_nota_base}, +{j_decimas}dec → {j_nota_final}')

# ── 2c. Agregar Julian Sandoval a Bienvenida ──────────────────
ws_bv = wb['Bienvenida']
last_r = 0
for row in ws_bv.iter_rows():
    for cell in row:
        if cell.value is not None:
            last_r = max(last_r, cell.row)

new_r = last_r + 1
ws_bv.cell(row=new_r, column=1).value = 'Julián Alonso Sandoval Arriagada'
ws_bv.cell(row=new_r, column=1).hyperlink = "#'Estudiante 23'!A1"
ws_bv.cell(row=new_r, column=1).font = Font(color='0563C1', underline='single')
print(f'  Bienvenida: Julián Sandoval agregado en fila {new_r} → Estudiante 23')

wb.save(XLSX)
print()
print('✅ feedback_estudiantes.xlsx guardado.')
print(f'   polar tv (Julián Aravena Sagal): 50pts → 4.0 base + 0.3 = 4.3 final')
print(f'   Julian Sandoval: 53pts → 4.2 base + 0.3 = 4.5 final')
