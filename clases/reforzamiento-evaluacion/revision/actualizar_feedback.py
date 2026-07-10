import openpyxl
import sys
from openpyxl.styles import Font, PatternFill

sys.stdout = open(sys.stdout.fileno(), mode='w', encoding='utf-8', buffering=1)

XLSX = r'C:/Users/diego/OneDrive/Escritorio/claude_codex/clases/claude_python/clases-python-4tomedio/clases/reforzamiento-evaluacion/revision/feedback_estudiantes.xlsx'
wb = openpyxl.load_workbook(XLSX)

NAME_MAP = {
    'Alex': 'Alex Saravia Lara',
    'Benjamín Diaz': 'Benjamín Ignacio Díaz Silva',
    'Cristóbal Muñoz': 'Cristóbal Alonso Muñoz Cubillos',
    'Damián Silva': 'Damián Cristóbal Flores Silva',
    'Diego Cifuentes': 'Diego Alonso Cifuentes Tessada',
    'Diego Donoso': 'Diego Andrés Donoso Figueroa',
    'Diego Peña': 'Diego Antonio Peña Bustamante',
    'Diego Vargas': 'Diego Pablo Vargas Jaqui',
    'Eduardo Eleazarth Pacco Ríos': 'Eduardo Eleazarth Pacco Ríos',
    'Felipe Aravena': 'Felipe Aravena Cárdenas',
    'Felipe Roman Brito': 'Felipe Alejandro Román Brito',
    'francisca parra': 'Francisca Belén Parra Marínquez',
    'francisco vega': 'Francisco Ignacio Vega Sanhueza',
    'Héctor _': 'Héctor Mauricio Vergara Alarcón',
    'Lucas gaspar Valenzuela donoso': 'Lucas Gaspar Valenzuela Donoso',
    'Luckas Letelier': 'Luckas Martín Letelier Lavín',
    'Maura Isabel Muñoz Gutierrez': 'Maura Isabel Muñoz Gutiérrez',
    'polar tv': 'Julián Rafael Aravena Sagal',
    'Sebastian Ulloa': 'Sebastián Andrés Ulloa Cuevas',
    'Vicente Benítez': 'Vicente Ignacio Benítez Muñoz',
    'Vicho 11': 'Vicente Narváez Fernandez',
}

DECIMAS = {
    'Alex Saravia Lara': 1,
    'Benjamín Ignacio Díaz Silva': 1,
    'Cristóbal Alonso Muñoz Cubillos': 3,
    'Damián Cristóbal Flores Silva': 3,
    'Diego Alonso Cifuentes Tessada': 0,
    'Diego Andrés Donoso Figueroa': 0,
    'Diego Antonio Peña Bustamante': 3,
    'Diego Pablo Vargas Jaqui': 1,
    'Eduardo Eleazarth Pacco Ríos': 3,
    'Felipe Aravena Cárdenas': 3,
    'Felipe Alejandro Román Brito': 3,
    'Francisca Belén Parra Marínquez': 0,
    'Francisco Ignacio Vega Sanhueza': 3,
    'Héctor Mauricio Vergara Alarcón': 3,
    'Lucas Gaspar Valenzuela Donoso': 1,
    'Luckas Martín Letelier Lavín': 3,
    'Maura Isabel Muñoz Gutiérrez': 1,
    'Julián Rafael Aravena Sagal': 3,
    'Sebastián Andrés Ulloa Cuevas': 1,
    'Vicente Ignacio Benítez Muñoz': 1,
    'Vicente Narváez Fernandez': 1,
    'Víctor Alonso Olguín': 0,
}

def calc_nota(score):
    if score < 50:
        return round(2 + 2 * (score / 50), 1)
    else:
        return round(4 + 3 * ((score - 50) / 50), 1)

# ── PASO 1: Actualizar nombres en Bienvenida ──────────────────
ws_bv = wb['Bienvenida']
for row in ws_bv.iter_rows():
    for cell in row:
        if cell.value in NAME_MAP:
            old = cell.value
            cell.value = NAME_MAP[old]
            print(f'  Bienvenida: {old!r} → {cell.value!r}')

# ── PASO 2: Mapa hoja → nombre oficial ───────────────────────
sheet_to_name = {}
for row in ws_bv.iter_rows():
    for cell in row:
        if cell.value and cell.hyperlink:
            link = cell.hyperlink.target
            if 'Estudiante' in link:
                sheet_name = link.split('!')[0].strip("#'")
                sheet_to_name[sheet_name] = cell.value

print()
for sh, nm in sorted(sheet_to_name.items()):
    print(f'  {sh} → {nm}')

# ── PASO 3: Agregar Décimas y Nota final a cada hoja ─────────
print()
for sh_name, oficial_name in sheet_to_name.items():
    ws = wb[sh_name]
    decimas = DECIMAS.get(oficial_name, 0)

    score = None
    nota_row = None
    for r in ws.iter_rows():
        for cell in r:
            if cell.value == 'TOTAL':
                raw = ws.cell(row=cell.row, column=4).value
                if raw:
                    score = int(str(raw).split('/')[0].strip())
            if cell.value == 'Nota':
                nota_row = cell.row

    if score is None or nota_row is None:
        print(f'  {sh_name}: SKIP (no score o nota_row)')
        continue

    nota_base = calc_nota(score)
    nota_final = round(min(7.0, nota_base + decimas * 0.1), 1)

    # Find last row with content to avoid overwriting merged comment rows
    last_r = 0
    for r in ws.iter_rows():
        for cell in r:
            if cell.value is not None:
                last_r = max(last_r, cell.row)
    r_dec = last_r + 1
    r_fin = last_r + 2

    ws.cell(row=r_dec, column=1).value = 'Décimas'
    ws.cell(row=r_dec, column=4).value = decimas
    ws.cell(row=r_fin, column=1).value = 'Nota final'
    ws.cell(row=r_fin, column=4).value = nota_final

    ws.cell(row=r_dec, column=1).font = Font(bold=True)
    ws.cell(row=r_fin, column=1).font = Font(bold=True, size=12)
    ws.cell(row=r_fin, column=4).font = Font(bold=True, size=12)

    print(f'  {sh_name} ({oficial_name}): {score}pts → nota_base={nota_base}, +{decimas}dec → nota_final={nota_final}')

# ── PASO 4: Nueva hoja Estudiante 22 — Víctor Alonso Olguín ──
print()
print('=== Creando Estudiante 22 (Víctor Alonso Olguín) ===')

victor_scores = [
    ('Promedio de notas', 7, 10, 'Faltan 3 prints individuales (-2), promedio sin etiqueta (-1).'),
    ('Reporte feria escolar', 10, 10, 'Titulo, encabezado columnas, sep, total: todo correcto.'),
    ('Simulador ahorro mensual', 15, 15, 'Usa una sola variable saldo con += y -= en cada paso. Caso ideal.'),
    ('Calculadora de descuento', 15, 15, 'input float ok, descuento y precio final correctos, datos en output.'),
    ('Reparto ganancias dulces', 12, 15, 'Calculos correctos. Olvido prefijo f en f-strings: prints muestran texto literal en vez de valores (-3).'),
    ('Depuracion 5 errores', 15, 15, 'Los 5 errores corregidos.'),
    ('Registro semanal de gastos', 20, 20, '6 inputs, total, promedio, saldo y header --- RESUMEN SEMANAL --- presentes.'),
]
v_total = sum(s[1] for s in victor_scores)
v_nota_base = calc_nota(v_total)
v_decimas = DECIMAS['Víctor Alonso Olguín']
v_nota_final = round(min(7.0, v_nota_base + v_decimas * 0.1), 1)

ws_v = wb.copy_worksheet(wb['Estudiante 1'])
ws_v.title = 'Estudiante 22'

ws_v.cell(row=1, column=1).value = 'Evaluacion 1  |  Python - Clases 1 a 7'
ws_v.cell(row=2, column=1).value = 'Ejercicio'
ws_v.cell(row=2, column=2).value = 'Obtenido'
ws_v.cell(row=2, column=3).value = 'Maximo'
ws_v.cell(row=2, column=4).value = 'Observacion'

for i, (ej, obt, mx, obs) in enumerate(victor_scores, start=3):
    ws_v.cell(row=i, column=1).value = ej
    ws_v.cell(row=i, column=2).value = obt
    ws_v.cell(row=i, column=3).value = mx
    ws_v.cell(row=i, column=4).value = obs

ws_v.cell(row=10, column=1).value = None
ws_v.cell(row=11, column=1).value = 'TOTAL'
ws_v.cell(row=11, column=4).value = f'{v_total} / 100'
ws_v.cell(row=12, column=1).value = 'Nota'
ws_v.cell(row=12, column=4).value = v_nota_base
ws_v.cell(row=13, column=1).value = 'Décimas'
ws_v.cell(row=13, column=4).value = v_decimas
ws_v.cell(row=14, column=1).value = 'Nota final'
ws_v.cell(row=14, column=4).value = v_nota_final

ws_v.cell(row=11, column=1).font = Font(bold=True)
ws_v.cell(row=12, column=1).font = Font(bold=True)
ws_v.cell(row=13, column=1).font = Font(bold=True)
ws_v.cell(row=14, column=1).font = Font(bold=True, size=12)
ws_v.cell(row=14, column=4).font = Font(bold=True, size=12)

warning_fill = PatternFill(start_color='FFAA00', end_color='FFAA00', fill_type='solid')
ws_v.cell(row=16, column=1).value = '⚠️ SOSPECHA DE COPIA — revisar con cuidado antes de publicar nota'
ws_v.cell(row=16, column=1).font = Font(bold=True, color='CC0000', size=12)
ws_v.cell(row=16, column=1).fill = warning_fill
try:
    ws_v.merge_cells('A16:D16')
except Exception:
    pass

print(f'  total={v_total}, nota_base={v_nota_base}, decimas={v_decimas}, nota_final={v_nota_final}')

# ── PASO 5: Añadir Víctor a Bienvenida ───────────────────────
last_row = 4
for row in ws_bv.iter_rows():
    for cell in row:
        if cell.value:
            last_row = max(last_row, cell.row)

new_r = last_row + 1
ws_bv.cell(row=new_r, column=1).value = 'Víctor Alonso Olguín'
ws_bv.cell(row=new_r, column=1).hyperlink = "#'Estudiante 22'!A1"
ws_bv.cell(row=new_r, column=1).font = Font(color='0563C1', underline='single')

warn_r = new_r + 1
ws_bv.cell(row=warn_r, column=1).value = '  ⚠️ Sospecha de copia — verificar antes de publicar'
ws_bv.cell(row=warn_r, column=1).font = Font(color='CC0000', bold=True, italic=True)

print(f'  Víctor agregado en fila {new_r} de Bienvenida')

# ── GUARDAR ───────────────────────────────────────────────────
wb.save(XLSX)
print()
print('✅ feedback_estudiantes.xlsx actualizado correctamente.')
