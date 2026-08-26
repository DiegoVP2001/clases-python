# -*- coding: utf-8 -*-
"""
Resumen final consolidado de la Evaluacion 2 (Clase 19, Condicionales) --
combina la version del dia (22 calificados + 9 ausentes) con la version
Atrasada (hasta 5 estudiantes con justificacion: Alain, Cristobal, Maura,
Simon, Diego Cifuentes).

Pensado para poder correrse en cualquier momento del proceso sin romperse:
si a alguien de la Atrasada aun le falta rendir, o si el metodo de
combinacion de Simon (dia + atrasada) sigue sin definirse, esos casos
quedan marcados como "Pendiente" en vez de hacer fallar el script.

Fuente de verdad de puntajes: NUNCA este script -- lee tal cual
  - version-dia/revision/puntajes_evaluacion2.json
  - version-atrasada/revision/puntajes.json
  - version-dia/revision/decimas_ejercitacion.csv
No escribe sobre ninguno de esos archivos, solo produce el resumen.

Cuando Simon termine la Atrasada y Diego decida como combinar su nota,
basta con fijar METODO_COMBINACION_SIMON abajo (o pasarlo por --metodo-simon)
y volver a correr -- no hace falta editar nada mas.

Uso (desde la raiz del proyecto):
    python "clases/clase-19-evaluacion-condicionales/generar_resumen_final_clase19.py"
    python "clases/clase-19-evaluacion-condicionales/generar_resumen_final_clase19.py" --metodo-simon promedio
"""

import argparse
import csv
import json
import sys
from pathlib import Path

CLASE_DIR = Path(__file__).resolve().parent
RAIZ = CLASE_DIR.parents[1]
sys.path.insert(0, str(RAIZ))

from tools.review_eval.colab_devolucion import calcular_nota, _redondear  # noqa: E402

PUNTAJES_DIA = CLASE_DIR / "version-dia" / "revision" / "puntajes_evaluacion2.json"
PUNTAJES_ATRASADA = CLASE_DIR / "version-atrasada" / "revision" / "puntajes.json"
DECIMAS_CSV = CLASE_DIR / "version-dia" / "revision" / "decimas_ejercitacion.csv"
OUT_DIR = CLASE_DIR / "revision-consolidada"

EXIGENCIA = 0.5
MAXIMO_DIA = 100

# Metodo para combinar la nota de Simon Abrahams (tiene intento el dia
# original, 34/100, MAS la Atrasada completa, 60/100).
#   "promedio"       -> (nota_dia + nota_atrasada) / 2
#   "mejor"          -> max(nota_dia, nota_atrasada)
#   "solo_atrasada"  -> se descarta el intento del dia, cuenta solo la atrasada
# Diego confirmo 2026-08-18 "solo_atrasada": se verifico item por item (11/11)
# que la Atrasada iguala o supera al dia en cada uno, asi que "mejor" habria
# dado exactamente el mismo resultado (4,6) -- "promedio" (4,1) penalizaria
# por un intento del dia que no aporto nada mejor en ningun item.
METODO_COMBINACION_SIMON = "solo_atrasada"

# Nombre corto (como aparece en puntajes_evaluacion2.json) -> clave en
# version-atrasada/revision/puntajes.json. Los 5 son el grupo de "atrasados
# con justificacion" (ver Historial.md, entrada 2026-08-14).
ATRASADOS_MAP = {
    "Alain Moyano": "alain_moyano",
    "Cristóbal Muñoz": "cristobal_munoz",
    "Maura Muñoz": "maura_munoz",
    "Diego Cifuentes": "diego_cifuentes",
    "Simón Abrahams": "simon_abrahams",
}

# N° de nomina (referencia-estudiantes/lista-estudiantes.md), para ordenar el
# resumen igual que el registro oficial del curso. Katalina (N°13) no aparece
# porque ya no esta en puntajes_evaluacion2.json (fuera del curso).
NOMINA = {
    "Eduardo Pacco": 1, "Diego Vargas": 2, "Alex Saravia": 3, "Francisco Vega": 4,
    "Felipe Román": 5, "Julián Sandoval": 6, "Héctor Vergara": 7, "Benjamín Mejías": 8,
    "Lucas Valenzuela": 9, "Martín Sánchez": 10, "Benjamín Díaz": 11, "Vicente Benítez": 12,
    "Maura Muñoz": 14, "Vicente Narváez": 15, "Sebastián Ulloa": 16, "Diego Donoso": 17,
    "Felipe Aravena": 18, "Cristóbal Muñoz": 19, "Francisca Parra": 20, "Diego Peña": 21,
    "Tomas Diaz": 22, "Santino Garcia": 23, "Luckas Letelier": 24, "Simón Abrahams": 25,
    "Julián Aravena": 26, "Diego Cifuentes": 27, "Paula Inzunza Gálvez": 28, "Damián Flores": 29,
    "Alain Moyano": 30, "Víctor Olguín": 31, "Nacor Pardo": 32,
}

# Mismo mapeo que generar_colab_personalizado_evaluacion2.py -- decimas de la
# Ejercitacion de Clase 17, aplican solo a los 22 calificados el dia original
# (los de la Atrasada ya traen su nota_final con decimas incluidas).
NOMBRE_EN_NOMINA = {
    "Alex Saravia": "Alex Saravia Lara",
    "Benjamín Díaz": "Benjamín Ignacio Díaz Silva",
    "Benjamín Mejías": "Benjamín Esteban Mejías González",
    "Damián Flores": "Damián Cristóbal Flores Silva",
    "Diego Donoso": "Diego Andrés Donoso Figueroa",
    "Diego Peña": "Diego Antonio Peña Bustamante",
    "Diego Vargas": "Diego Pablo Vargas Jaqui",
    "Eduardo Pacco": "Eduardo Eleazarth Pacco Ríos",
    "Felipe Aravena": "Felipe Aravena Cárdenas",
    "Felipe Román": "Felipe Alejandro Román Brito",
    "Francisco Vega": "Francisco Ignacio Vega Sanhueza",
    "Héctor Vergara": "Héctor Mauricio Vergara Alarcón",
    "Julián Aravena": "Julián Rafael Aravena Sagal",
    "Julián Sandoval": "Julián Alonso Sandoval Arriagada",
    "Luckas Letelier": "Luckas Martín Letelier Lavín",
    "Martín Sánchez": "Martín Natahel Sánchez Orellana",
    "Santino Garcia": "Santino Garcia Colombati",
    "Sebastián Ulloa": "Sebastián Andrés Ulloa Cuevas",
    "Simón Abrahams": "Simón Abrahams Delgado",
    "Tomas Diaz": "Tomas Ignacio Diaz Calfunao",
    "Vicente Benítez": "Vicente Ignacio Benítez Muñoz",
    "Vicente Narváez": "Vicente Narváez Fernandez",
}


def cargar_decimas() -> dict:
    with DECIMAS_CSV.open(encoding="utf-8", newline="") as fh:
        por_nomina = {fila["Nombre completo"]: int(fila["Total"]) for fila in csv.DictReader(fh)}
    decimas = {}
    for corto, completo in NOMBRE_EN_NOMINA.items():
        assert completo in por_nomina, f"{corto}: '{completo}' no está en {DECIMAS_CSV.name}"
        decimas[corto] = por_nomina[completo]
    return decimas


def _coma(valor: float) -> str:
    return f"{valor:.1f}".replace(".", ",")


def combinar_simon(nota_dia: float, nota_atrasada: float, metodo: str) -> float:
    if metodo == "promedio":
        return _redondear((nota_dia + nota_atrasada) / 2)
    if metodo == "mejor":
        return max(nota_dia, nota_atrasada)
    if metodo == "solo_atrasada":
        return nota_atrasada
    raise ValueError(f"Método de combinación desconocido: {metodo!r}")


def build_rows(metodo_simon: str | None) -> list[dict]:
    dia = json.loads(PUNTAJES_DIA.read_text(encoding="utf-8"))["estudiantes"]
    atrasada_path_existe = PUNTAJES_ATRASADA.exists()
    atrasada = (
        json.loads(PUNTAJES_ATRASADA.read_text(encoding="utf-8"))["estudiantes"]
        if atrasada_path_existe else {}
    )
    decimas = cargar_decimas()

    assert "Katalina Domínguez Santibáñez" not in dia, (
        "Katalina ya no está en el curso -- no debería aparecer en puntajes_evaluacion2.json"
    )

    rows = []
    for nombre, info in dia.items():
        n_nomina = NOMINA.get(nombre)
        assert n_nomina is not None, f"'{nombre}' no está en el diccionario NOMINA de este script"

        es_atrasado = nombre in ATRASADOS_MAP
        total_dia = info.get("total") or 0
        rindio_dia = bool(info.get("revisado")) and total_dia > 0

        nota_dia = None
        if rindio_dia:
            nota_base = calcular_nota(total_dia, MAXIMO_DIA, EXIGENCIA)
            nota_dia = min(7.0, _redondear(nota_base + decimas.get(nombre, 0) / 10))

        nota_atrasada = None
        total_atrasada = None
        rindio_atrasada = False
        if es_atrasado:
            clave = ATRASADOS_MAP[nombre]
            reg = atrasada.get(clave)
            if reg is not None:
                rindio_atrasada = True
                nota_atrasada = reg["nota_final"]
                total_atrasada = reg["total"]

        # --- Decidir nota final + estado ------------------------------------
        if es_atrasado and rindio_dia and rindio_atrasada:
            # Caso Simón: dos instancias completas, requiere método de combinación.
            if metodo_simon:
                nota_final = combinar_simon(nota_dia, nota_atrasada, metodo_simon)
                fuente = f"Día+Atrasada ({metodo_simon})"
                estado = "Aprobado" if nota_final >= 4.0 else "Reprobado"
            else:
                nota_final = None
                fuente = f"Día {_coma(nota_dia)} + Atrasada {_coma(nota_atrasada)}"
                estado = "PENDIENTE — falta definir método de combinación"
        elif es_atrasado and rindio_atrasada:
            nota_final = nota_atrasada
            fuente = "Atrasada"
            estado = "Aprobado" if nota_final >= 4.0 else "Reprobado"
        elif es_atrasado and not rindio_atrasada:
            nota_final = None
            fuente = "—"
            estado = "Pendiente — no ha rendido la Atrasada"
        elif rindio_dia:
            nota_final = nota_dia
            fuente = "Día"
            estado = "Aprobado" if nota_final >= 4.0 else "Reprobado"
        else:
            nota_final = None
            fuente = "—"
            estado = "Ausente"

        rows.append({
            "N°": n_nomina,
            "Nombre": nombre,
            "Puntaje día": total_dia if rindio_dia else None,
            "Puntaje atrasada": total_atrasada,
            "Fuente": fuente,
            "Nota": nota_final,
            "Estado": estado,
        })

    rows.sort(key=lambda r: r["N°"])
    return rows


def generar_md(rows: list[dict]) -> None:
    con_nota = [r for r in rows if r["Nota"] is not None]
    pendientes = [r for r in rows if r["Estado"].startswith("Pendiente") or r["Estado"].startswith("PENDIENTE")]
    ausentes = [r for r in rows if r["Estado"] == "Ausente"]

    lines = [
        "# Resumen Final — Clase 19 (Condicionales) — Día + Atrasada",
        "",
        f"Total en registro: {len(rows)} | Con nota: {len(con_nota)} | "
        f"Pendientes: {len(pendientes)} | Ausentes: {len(ausentes)}",
        "",
        "## Tabla de notas",
        "",
        "| N° | Nombre | Pje. día | Pje. atrasada | Fuente | Nota | Estado |",
        "| --- | --- | --- | --- | --- | --- | --- |",
    ]
    for r in rows:
        lines.append(
            f"| {r['N°']} | {r['Nombre']} | {r['Puntaje día'] if r['Puntaje día'] is not None else '—'} | "
            f"{r['Puntaje atrasada'] if r['Puntaje atrasada'] is not None else '—'} | {r['Fuente']} | "
            f"{_coma(r['Nota']) if r['Nota'] is not None else '—'} | {r['Estado']} |"
        )

    if con_nota:
        notas = [r["Nota"] for r in con_nota]
        aprobados = sum(1 for n in notas if n >= 4.0)
        lines += [
            "",
            "## Estadísticas (solo estudiantes con nota)",
            "",
            "| Métrica | Valor |",
            "| --- | --- |",
            f"| Con nota | {len(con_nota)} |",
            f"| Aprobados | {aprobados} / {len(con_nota)} ({aprobados/len(con_nota)*100:.0f}%) |",
            f"| Promedio | {_coma(sum(notas)/len(notas))} |",
            f"| Máximo | {_coma(max(notas))} |",
            f"| Mínimo | {_coma(min(notas))} |",
        ]

    if pendientes:
        lines += ["", "## Pendientes", ""]
        for r in pendientes:
            lines.append(f"- **{r['Nombre']}** (N°{r['N°']}): {r['Estado']}")

    if ausentes:
        lines += ["", "## Ausentes (no rindieron ninguna versión)", ""]
        for r in ausentes:
            lines.append(f"- {r['Nombre']} (N°{r['N°']})")

    OUT_DIR.mkdir(exist_ok=True)
    (OUT_DIR / "resumen_final_clase19.md").write_text("\n".join(lines), encoding="utf-8")
    print("  OK  resumen_final_clase19.md")


def generar_csv(rows: list[dict]) -> None:
    OUT_DIR.mkdir(exist_ok=True)
    path = OUT_DIR / "resumen_final_clase19.csv"
    campos = ["N°", "Nombre", "Puntaje día", "Puntaje atrasada", "Fuente", "Nota", "Estado"]
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=campos)
        writer.writeheader()
        for r in rows:
            fila = dict(r)
            fila["Nota"] = _coma(r["Nota"]) if r["Nota"] is not None else ""
            writer.writerow(fila)
    print("  OK  resumen_final_clase19.csv")


def generar_xlsx(rows: list[dict]) -> None:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    COLOR_HEADER = "1F4E79"
    COLOR_APROBADO = "E2EFDA"
    COLOR_REPROBADO = "FFDCDC"
    COLOR_PENDIENTE = "FFF2CC"
    COLOR_AUSENTE = "EDEDED"
    WHITE = "FFFFFF"
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resumen Clase 19"

    headers = ["N°", "Nombre", "Pje. día", "Pje. atrasada", "Fuente", "Nota", "Estado"]
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title = ws.cell(row=1, column=1, value="Clase 19 — Condicionales — Resumen Final (Día + Atrasada)")
    title.font = Font(bold=True, color=WHITE, size=13)
    title.fill = PatternFill("solid", fgColor=COLOR_HEADER)
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=2, column=col, value=h)
        c.font = Font(bold=True, color=WHITE)
        c.fill = PatternFill("solid", fgColor=COLOR_HEADER)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border

    for i, r in enumerate(rows):
        row_i = i + 3
        if r["Estado"] == "Aprobado":
            fill = PatternFill("solid", fgColor=COLOR_APROBADO)
        elif r["Estado"] == "Reprobado":
            fill = PatternFill("solid", fgColor=COLOR_REPROBADO)
        elif r["Estado"] == "Ausente":
            fill = PatternFill("solid", fgColor=COLOR_AUSENTE)
        else:
            fill = PatternFill("solid", fgColor=COLOR_PENDIENTE)

        valores = [
            r["N°"], r["Nombre"],
            r["Puntaje día"] if r["Puntaje día"] is not None else "—",
            r["Puntaje atrasada"] if r["Puntaje atrasada"] is not None else "—",
            r["Fuente"],
            _coma(r["Nota"]) if r["Nota"] is not None else "—",
            r["Estado"],
        ]
        for col, val in enumerate(valores, start=1):
            c = ws.cell(row=row_i, column=col, value=val)
            c.border = border
            c.fill = fill
            c.alignment = Alignment(horizontal="left" if col in (2, 7) else "center", vertical="center", wrap_text=True)

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 8
    ws.column_dimensions["G"].width = 42
    ws.freeze_panes = "A3"

    OUT_DIR.mkdir(exist_ok=True)
    path = OUT_DIR / "resumen_final_clase19.xlsx"
    wb.save(path)
    print("  OK  resumen_final_clase19.xlsx")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--metodo-simon",
        choices=["promedio", "mejor", "solo_atrasada"],
        default=METODO_COMBINACION_SIMON,
        help="Cómo combinar la nota de Simón (día + atrasada). Sin definir, queda como Pendiente.",
    )
    args = parser.parse_args()

    rows = build_rows(args.metodo_simon)
    print(f"\nGenerando resumen final de Clase 19 ({len(rows)} estudiantes en registro)...\n")
    generar_md(rows)
    generar_csv(rows)
    generar_xlsx(rows)
    print(f"\nListo. Archivos en {OUT_DIR.relative_to(RAIZ)}/")


if __name__ == "__main__":
    main()
